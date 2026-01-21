"""
Excel Export Service for Reports

Provides multi-sheet Excel export functionality for all report pages.
Supports:
- Multi-tenant data isolation
- Filter-aware exports
- Multiple sheets per export
- Proper data typing (dates, decimals, numbers)
- Performance optimization for large datasets

Dependencies: openpyxl
"""

import io
from datetime import datetime, date
from decimal import Decimal
from typing import List, Dict, Any, Optional, Union
from dataclasses import dataclass

from django.db.models import Sum, Count, Avg, F, DecimalField, Q
from django.db.models.functions import Coalesce, TruncDate
from django.http import HttpResponse
from django.utils import timezone

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
    
    # Styling constants - only defined if openpyxl is available
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    CELL_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
except ImportError:
    OPENPYXL_AVAILABLE = False
    # Define dummy constants to prevent NameError
    HEADER_FONT = None
    HEADER_FILL = None
    HEADER_ALIGNMENT = None
    CELL_ALIGNMENT = None
    THIN_BORDER = None

from orders.models import Order
from accounts.models import BotUser
from organizations.models import Branch, TranslationCenter, AdminUser
from organizations.rbac import get_user_orders, get_user_customers, get_user_branches

# Status display mappings
STATUS_LABELS = {
    "pending": "Pending",
    "payment_pending": "Awaiting",
    "payment_received": "Received",
    "payment_confirmed": "Confirmed",
    "in_progress": "In Process",
    "ready": "Ready",
    "completed": "Done",
    "cancelled": "Cancelled",
}

PAYMENT_TYPE_LABELS = {
    "cash": "Cash",
    "card": "Card",
    "transfer": "Transfer",
    "": "Not Set",
    None: "Not Set",
}


@dataclass
class SheetConfig:
    """Configuration for a single Excel sheet"""
    name: str
    headers: List[str]
    data: List[List[Any]]
    column_widths: Optional[List[int]] = None


class ExcelExporter:
    """
    Main Excel export class with multi-sheet support.
    
    Usage:
        exporter = ExcelExporter()
        exporter.add_sheet(SheetConfig(...))
        response = exporter.generate_response("filename.xlsx")
    """
    
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        self.workbook = Workbook()
        # Remove default sheet
        self.workbook.remove(self.workbook.active)
        self.sheets: List[SheetConfig] = []
    
    def add_sheet(self, config: SheetConfig):
        """Add a sheet configuration to be rendered"""
        self.sheets.append(config)
    
    def _format_cell_value(self, value: Any) -> Any:
        """Format cell value to appropriate Excel type"""
        if value is None:
            return ""
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, datetime):
            return value.replace(tzinfo=None) if value.tzinfo else value
        if isinstance(value, date):
            return value
        if isinstance(value, bool):
            return "Yes" if value else "No"
        return value
    
    def _auto_column_width(self, ws, col_idx: int, values: List[Any], header: str) -> int:
        """Calculate auto column width based on content"""
        max_length = len(str(header))
        for value in values[:100]:  # Check first 100 rows for performance
            cell_length = len(str(value)) if value else 0
            max_length = max(max_length, cell_length)
        # Add padding and cap at reasonable max
        return min(max(max_length + 2, 10), 50)
    
    def _render_sheet(self, config: SheetConfig):
        """Render a single sheet from configuration"""
        ws = self.workbook.create_sheet(title=config.name[:31])  # Excel sheet name limit
        
        # Write headers
        for col_idx, header in enumerate(config.headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER
        
        # Write data rows
        for row_idx, row_data in enumerate(config.data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=self._format_cell_value(value))
                cell.alignment = CELL_ALIGNMENT
                cell.border = THIN_BORDER
                
                # Apply number format for decimals/money
                if isinstance(value, (Decimal, float)) and col_idx > 0:
                    header_lower = config.headers[col_idx - 1].lower()
                    if any(x in header_lower for x in ['price', 'revenue', 'amount', 'sum', 'total', 'paid', 'remaining']):
                        cell.number_format = '#,##0.00'
                    elif 'rate' in header_lower or '%' in header_lower:
                        cell.number_format = '0.0%' if value < 1 else '0.0'
        
        # Set column widths
        if config.column_widths:
            for col_idx, width in enumerate(config.column_widths, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width
        else:
            # Auto-calculate widths
            for col_idx, header in enumerate(config.headers, 1):
                col_values = [row[col_idx - 1] if len(row) > col_idx - 1 else "" for row in config.data]
                width = self._auto_column_width(ws, col_idx, col_values, header)
                ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Freeze header row
        ws.freeze_panes = "A2"
    
    def render(self):
        """Render all sheets"""
        for config in self.sheets:
            self._render_sheet(config)
    
    def generate_response(self, filename: str) -> HttpResponse:
        """Generate HTTP response with Excel file"""
        self.render()
        
        # Write to buffer
        buffer = io.BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        
        # Create response
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def is_empty(self) -> bool:
        """Check if all sheets are empty"""
        return all(len(config.data) == 0 for config in self.sheets)
    
    def to_bytes(self) -> bytes:
        """Generate Excel file and return as bytes"""
        self.render()
        buffer = io.BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


class ReportExporter:
    """
    Main Report Export facade class.
    
    Wraps all individual report export functions and provides a unified interface
    for exporting any report type with filters.
    
    Usage:
        exporter = ReportExporter(request.user)
        result = exporter.export('orders', {'period': 'month', 'branch_id': 1})
        
        if result['success']:
            # result['file_content'] is the Excel bytes
            # result['filename'] is the suggested filename
        else:
            # result['message'] contains error/warning message
    """
    
    def __init__(self, user):
        self.user = user
    
    def _parse_date_filters(self, filters: Dict) -> tuple:
        """Parse date range from filter parameters"""
        from datetime import timedelta
        
        period = filters.get('period', 'month')
        custom_from = filters.get('date_from')
        custom_to = filters.get('date_to')
        
        today = timezone.now()
        
        if period == 'today':
            date_from = today.replace(hour=0, minute=0, second=0, microsecond=0)
            date_to = today
        elif period == 'yesterday':
            yesterday = today - timedelta(days=1)
            date_from = yesterday.replace(hour=0, minute=0, second=0, microsecond=0)
            date_to = yesterday.replace(hour=23, minute=59, second=59, microsecond=999999)
        elif period == 'week':
            start_of_week = today - timedelta(days=today.weekday())
            date_from = start_of_week.replace(hour=0, minute=0, second=0, microsecond=0)
            date_to = today
        elif period == 'quarter':
            quarter = (today.month - 1) // 3
            start_month = quarter * 3 + 1
            date_from = today.replace(month=start_month, day=1, hour=0, minute=0, second=0, microsecond=0)
            date_to = today
        elif period == 'year':
            date_from = today.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            date_to = today
        elif period == 'custom' and custom_from and custom_to:
            try:
                if isinstance(custom_from, str):
                    date_from = datetime.strptime(custom_from, '%Y-%m-%d')
                    date_from = timezone.make_aware(date_from.replace(hour=0, minute=0, second=0))
                else:
                    date_from = custom_from
                if isinstance(custom_to, str):
                    date_to = datetime.strptime(custom_to, '%Y-%m-%d')
                    date_to = timezone.make_aware(date_to.replace(hour=23, minute=59, second=59))
                else:
                    date_to = custom_to
            except ValueError:
                # Fallback to month
                date_from = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                date_to = today
        else:
            # Default: month
            date_from = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            date_to = today
        
        return date_from, date_to
    
    def export(self, report_type: str, filters: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        Export the specified report type with given filters.
        
        Args:
            report_type: One of 'orders', 'financial', 'staff_performance', 
                        'branch_comparison', 'customers', 'unit_economy', 'my_statistics'
            filters: Dict with filter parameters (period, date_from, date_to, branch_id, center_id, etc.)
        
        Returns:
            Dict with keys:
                - success: bool
                - file_content: bytes (if success)
                - filename: str (if success)
                - message: str (error/warning message)
        """
        if filters is None:
            filters = {}
        
        try:
            # Parse date range
            date_from, date_to = self._parse_date_filters(filters)
            
            # Extract common filters
            branch_id = filters.get('branch_id')
            center_id = filters.get('center_id')
            status_filter = filters.get('status')
            
            # Call appropriate export function
            if report_type == 'orders':
                response = export_orders_report(
                    self.user, date_from, date_to, branch_id, center_id, status_filter
                )
            elif report_type == 'financial':
                response = export_financial_report(
                    self.user, date_from, date_to, branch_id, center_id
                )
            elif report_type == 'staff_performance':
                response = export_staff_performance(
                    self.user, date_from, date_to, branch_id, center_id
                )
            elif report_type == 'branch_comparison':
                response = export_branch_comparison(
                    self.user, date_from, date_to, center_id
                )
            elif report_type == 'customers':
                response = export_customer_analytics(
                    self.user, date_from, date_to, branch_id, center_id
                )
            elif report_type == 'unit_economy':
                response = export_unit_economy(
                    self.user, branch_id, center_id
                )
            elif report_type == 'my_statistics':
                response = export_my_statistics(
                    self.user, date_from, date_to
                )
            elif report_type == 'expense_analytics':
                expense_type_filter = filters.get('expense_type')
                response = export_expense_analytics(
                    self.user, date_from, date_to, branch_id, center_id, expense_type_filter
                )
            else:
                return {
                    'success': False,
                    'message': f'Unknown report type: {report_type}'
                }
            
            # Extract file content and filename from HttpResponse
            file_content = response.content
            filename = response['Content-Disposition'].split('filename="')[1].rstrip('"')
            
            # Check if empty (no data)
            if len(file_content) < 1000:  # Very small file likely means empty
                return {
                    'success': False,
                    'message': 'No data found for the selected filters. Try adjusting your date range or filters.'
                }
            
            return {
                'success': True,
                'file_content': file_content,
                'filename': filename,
                'message': 'Export generated successfully'
            }
            
        except ImportError as e:
            return {
                'success': False,
                'message': 'Excel export is not available. Please install openpyxl: pip install openpyxl'
            }
        except Exception as e:
            import logging
            logging.getLogger(__name__).exception(f"Export error for {report_type}")
            return {
                'success': False,
                'message': f'Export failed: {str(e)}'
            }


# ============ Report-Specific Export Functions ============

def export_orders_report(
    user,
    date_from: datetime,
    date_to: datetime,
    branch_id: Optional[str] = None,
    center_id: Optional[str] = None,
    status_filter: Optional[str] = None,
) -> HttpResponse:
    """
    Export Orders Report with multiple sheets:
    - Executive Summary
    - Orders (Detailed)
    - Status Breakdown
    - Language Breakdown
    - Product Analytics
    - Customer Type Breakdown
    - Daily Summary
    """
    exporter = ExcelExporter()
    
    # Get filtered orders
    orders = get_user_orders(user).filter(
        created_at__gte=date_from,
        created_at__lte=date_to
    ).select_related('bot_user', 'product', 'product__category', 'branch', 'assigned_to__user', 'language')
    
    if center_id:
        orders = orders.filter(branch__center_id=center_id)
    if branch_id:
        orders = orders.filter(branch_id=branch_id)
    if status_filter:
        orders = orders.filter(status=status_filter)
    
    # Calculate overall statistics
    total_orders = orders.count()
    total_revenue = orders.aggregate(total=Sum('total_price'))['total'] or 0
    total_received = orders.aggregate(total=Sum('received'))['total'] or 0
    total_pages = orders.aggregate(total=Sum('total_pages'))['total'] or 0
    avg_order = orders.aggregate(avg=Avg('total_price'))['avg'] or 0
    completed = orders.filter(status='completed').count()
    b2b_orders = orders.filter(bot_user__is_agency=True).count()
    b2c_orders = orders.filter(bot_user__is_agency=False).count()
    unique_customers = orders.values('bot_user').distinct().count()
    
    # Sheet 1: Executive Summary
    summary_data = [
        ["REPORT INFORMATION", ""],
        ["Report Type", "Orders Report"],
        ["Generated Date", datetime.now().strftime('%Y-%m-%d %H:%M')],
        ["Period From", date_from.strftime('%Y-%m-%d')],
        ["Period To", date_to.strftime('%Y-%m-%d')],
        ["Generated By", user.get_full_name() or user.username],
        ["", ""],
        ["ORDER OVERVIEW", ""],
        ["Total Orders", total_orders],
        ["Completed Orders", completed],
        ["Completion Rate", f"{(completed/total_orders*100):.1f}%" if total_orders > 0 else "0%"],
        ["Total Pages Processed", total_pages or 0],
        ["Unique Customers", unique_customers],
        ["", ""],
        ["FINANCIAL SUMMARY", ""],
        ["Total Revenue", float(total_revenue)],
        ["Total Received", float(total_received)],
        ["Outstanding Balance", float(total_revenue - total_received)],
        ["Collection Rate", f"{(total_received/total_revenue*100):.1f}%" if total_revenue > 0 else "0%"],
        ["Average Order Value", float(avg_order)],
        ["", ""],
        ["CUSTOMER BREAKDOWN", ""],
        ["B2B Orders (Agencies)", b2b_orders],
        ["B2B Percentage", f"{(b2b_orders/total_orders*100):.1f}%" if total_orders > 0 else "0%"],
        ["B2C Orders (Regular)", b2c_orders],
        ["B2C Percentage", f"{(b2c_orders/total_orders*100):.1f}%" if total_orders > 0 else "0%"],
    ]
    
    exporter.add_sheet(SheetConfig(
        name="Executive Summary",
        headers=["Metric", "Value"],
        data=summary_data,
        column_widths=[35, 25]
    ))
    
    # Sheet 2: Detailed Orders
    orders_data = []
    for order in orders.order_by('-created_at'):
        orders_data.append([
            order.id,
            order.created_at,
            order.bot_user.name if order.bot_user else "N/A",
            order.bot_user.phone if order.bot_user else "N/A",
            "B2B" if (order.bot_user and order.bot_user.is_agency) else "B2C",
            order.product.name if order.product else "N/A",
            order.product.category.name if (order.product and order.product.category) else "N/A",
            order.language.name if order.language else "N/A",
            order.total_pages or 0,
            order.copy_number or 0,
            order.total_price or 0,
            order.received or 0,
            (order.total_price or 0) - (order.received or 0),
            STATUS_LABELS.get(order.status, order.status),
            PAYMENT_TYPE_LABELS.get(order.payment_type, order.payment_type or "N/A"),
            order.branch.name if order.branch else "N/A",
            order.assigned_to.user.get_full_name() if order.assigned_to else "Unassigned",
            order.completed_at if order.completed_at else None,
        ])
    
    exporter.add_sheet(SheetConfig(
        name="Orders (Detailed)",
        headers=[
            "Order ID", "Created At", "Customer Name", "Phone", "Customer Type",
            "Product", "Category", "Language", "Pages", "Copies",
            "Total Price", "Paid Amount", "Remaining", "Status", "Payment Type",
            "Branch", "Assigned To", "Completed At"
        ],
        data=orders_data
    ))
    
    # Sheet 2: Status Breakdown
    status_breakdown = orders.values('status').annotate(
        count=Count('id'),
        revenue=Coalesce(Sum('total_price'), Decimal('0'), output_field=DecimalField()),
        paid=Coalesce(Sum('received'), Decimal('0'), output_field=DecimalField())
    ).order_by('-count')
    
    status_data = []
    for item in status_breakdown:
        status_data.append([
            STATUS_LABELS.get(item['status'], item['status']),
            item['count'],
            float(item['revenue'] or 0),
            float(item['paid'] or 0),
            float((item['revenue'] or 0) - (item['paid'] or 0)),
        ])
    
    exporter.add_sheet(SheetConfig(
        name="Status Breakdown",
        headers=["Status", "Orders Count", "Total Revenue", "Paid Amount", "Outstanding"],
        data=status_data
    ))
    
    # Sheet 3: Language Breakdown
    lang_breakdown = orders.values('language__name').annotate(
        count=Count('id'),
        revenue=Coalesce(Sum('total_price'), Decimal('0'), output_field=DecimalField())
    ).order_by('-count')
    
    lang_data = []
    for item in lang_breakdown:
        lang_data.append([
            item['language__name'] or "Not Specified",
            item['count'],
            float(item['revenue'] or 0),
        ])
    
    exporter.add_sheet(SheetConfig(
        name="Language Breakdown",
        headers=["Language", "Orders Count", "Revenue"],
        data=lang_data
    ))
    
    # Sheet 4: Product Analytics
    product_analytics = orders.values('product__name', 'product__category__name').annotate(
        count=Count('id'),
        revenue=Coalesce(Sum('total_price'), Decimal('0'), output_field=DecimalField()),
        pages=Coalesce(Sum('total_pages'), 0),
        avg_price=Coalesce(Avg('total_price'), Decimal('0'), output_field=DecimalField())
    ).order_by('-revenue')
    
    product_data = []
    for item in product_analytics:
        pct = (item['revenue'] / total_revenue * 100) if total_revenue > 0 else 0
        product_data.append([
            item['product__name'] or "Unknown",
            item['product__category__name'] or "N/A",
            item['count'],
            f"{pct:.1f}%",
            item['pages'] or 0,
            float(item['revenue'] or 0),
            float(item['avg_price'] or 0),
        ])
    
    exporter.add_sheet(SheetConfig(
        name="Product Analytics",
        headers=["Product", "Category", "Orders", "% of Total", "Pages", "Revenue", "Avg Price"],
        data=product_data,
        column_widths=[25, 20, 10, 12, 10, 15, 15]
    ))
    
    # Sheet 5: Customer Type Breakdown
    b2b_revenue = orders.filter(bot_user__is_agency=True).aggregate(total=Sum('total_price'))['total'] or 0
    b2c_revenue = orders.filter(bot_user__is_agency=False).aggregate(total=Sum('total_price'))['total'] or 0
    b2b_pages = orders.filter(bot_user__is_agency=True).aggregate(total=Sum('total_pages'))['total'] or 0
    b2c_pages = orders.filter(bot_user__is_agency=False).aggregate(total=Sum('total_pages'))['total'] or 0
    
    customer_type_data = [
        [
            "B2B (Agencies)",
            b2b_orders,
            f"{(b2b_orders/total_orders*100):.1f}%" if total_orders > 0 else "0%",
            b2b_pages or 0,
            float(b2b_revenue),
            f"{(b2b_revenue/total_revenue*100):.1f}%" if total_revenue > 0 else "0%",
            float(b2b_revenue / b2b_orders) if b2b_orders > 0 else 0,
        ],
        [
            "B2C (Regular Customers)",
            b2c_orders,
            f"{(b2c_orders/total_orders*100):.1f}%" if total_orders > 0 else "0%",
            b2c_pages or 0,
            float(b2c_revenue),
            f"{(b2c_revenue/total_revenue*100):.1f}%" if total_revenue > 0 else "0%",
            float(b2c_revenue / b2c_orders) if b2c_orders > 0 else 0,
        ],
    ]
    
    exporter.add_sheet(SheetConfig(
        name="Customer Type Analysis",
        headers=["Customer Type", "Orders", "% Orders", "Pages", "Revenue", "% Revenue", "Avg Order Value"],
        data=customer_type_data,
        column_widths=[25, 10, 12, 10, 15, 12, 18]
    ))
    
    # Sheet 6: Daily Summary
    daily_breakdown = orders.annotate(
        date=TruncDate('created_at')
    ).values('date').annotate(
        count=Count('id'),
        revenue=Coalesce(Sum('total_price'), Decimal('0'), output_field=DecimalField()),
        pages=Coalesce(Sum('total_pages'), 0),
        paid=Coalesce(Sum('received'), Decimal('0'), output_field=DecimalField())
    ).order_by('date')
    
    daily_data = []
    for item in daily_breakdown:
        daily_data.append([
            item['date'],
            item['count'],
            item['pages'] or 0,
            float(item['revenue'] or 0),
            float(item['paid'] or 0),
        ])
    
    exporter.add_sheet(SheetConfig(
        name="Daily Summary",
        headers=["Date", "Orders", "Total Pages", "Revenue", "Paid"],
        data=daily_data
    ))
    
    # Generate filename
    center_name = ""
    if center_id:
        try:
            center_name = f"_{TranslationCenter.objects.get(id=center_id).name}"
        except:
            pass
    
    filename = f"orders_report{center_name}_{date_from.strftime('%Y%m%d')}_to_{date_to.strftime('%Y%m%d')}.xlsx"
    
    return exporter.generate_response(filename)


def export_financial_report(
    user,
    date_from: datetime,
    date_to: datetime,
    branch_id: Optional[str] = None,
    center_id: Optional[str] = None,
) -> HttpResponse:
    """
    Export Financial Report with multiple sheets:
    - Executive Summary
    - Revenue by Status
    - Revenue by Branch
    - Revenue by Product
    - Payment Method Breakdown
    - Daily Revenue Trend
    - Payment Details
    """
    exporter = ExcelExporter()
    
    # Get filtered orders
    orders = get_user_orders(user).filter(
        created_at__gte=date_from,
        created_at__lte=date_to
    ).select_related('bot_user', 'product', 'branch')
    
    if center_id:
        orders = orders.filter(branch__center_id=center_id)
    if branch_id:
        orders = orders.filter(branch_id=branch_id)
    
    # Calculate statistics
    total_revenue = orders.aggregate(total=Sum('total_price'))['total'] or 0
    total_received = orders.aggregate(total=Sum('received'))['total'] or 0
    total_orders = orders.count()
    avg_order = orders.aggregate(avg=Avg('total_price'))['avg'] or 0
    completed_orders = orders.filter(status='completed').count()
    completed_revenue = orders.filter(status='completed').aggregate(total=Sum('total_price'))['total'] or 0
    total_pages = orders.aggregate(total=Sum('total_pages'))['total'] or 0
    b2b_revenue = orders.filter(bot_user__is_agency=True).aggregate(total=Sum('total_price'))['total'] or 0
    b2c_revenue = orders.filter(bot_user__is_agency=False).aggregate(total=Sum('total_price'))['total'] or 0
    
    # Sheet 1: Executive Summary
    summary_data = [
        ["REPORT INFORMATION", ""],
        ["Report Type", "Financial Report"],
        ["Generated Date", datetime.now().strftime('%Y-%m-%d %H:%M')],
        ["Period From", date_from.strftime('%Y-%m-%d')],
        ["Period To", date_to.strftime('%Y-%m-%d')],
        ["Generated By", user.get_full_name() or user.username],
        ["", ""],
        ["REVENUE OVERVIEW", ""],
        ["Total Orders", total_orders],
        ["Total Revenue", float(total_revenue)],
        ["Total Received", float(total_received)],
        ["Outstanding Balance", float(total_revenue - total_received)],
        ["Collection Rate", f"{(total_received/total_revenue*100):.1f}%" if total_revenue > 0 else "0%"],
        ["Average Order Value", float(avg_order)],
        ["Revenue per Page", float(total_revenue / total_pages) if total_pages > 0 else 0],
        ["", ""],
        ["ORDER STATUS", ""],
        ["Completed Orders", completed_orders],
        ["Completion Rate", f"{(completed_orders/total_orders*100):.1f}%" if total_orders > 0 else "0%"],
        ["Completed Revenue", float(completed_revenue)],
        ["Pending Revenue", float(total_revenue - completed_revenue)],
        ["", ""],
        ["CUSTOMER TYPE REVENUE", ""],
        ["B2B Revenue", float(b2b_revenue)],
        ["B2B Percentage", f"{(b2b_revenue/total_revenue*100):.1f}%" if total_revenue > 0 else "0%"],
        ["B2C Revenue", float(b2c_revenue)],
        ["B2C Percentage", f"{(b2c_revenue/total_revenue*100):.1f}%" if total_revenue > 0 else "0%"],
    ]
    
    exporter.add_sheet(SheetConfig(
        name="Executive Summary",
        headers=["Metric", "Value"],
        data=summary_data,
        column_widths=[35, 25]
    ))
    
    # Sheet 2: Revenue by Status
    status_revenue = orders.values('status').annotate(
        count=Count('id'),
        revenue=Coalesce(Sum('total_price'), Decimal('0'), output_field=DecimalField()),
        received=Coalesce(Sum('received'), Decimal('0'), output_field=DecimalField())
    ).order_by('-revenue')
    
    status_data = []
    for item in status_revenue:
        status_data.append([
            STATUS_LABELS.get(item['status'], item['status']),
            item['count'],
            float(item['revenue'] or 0),
            float(item['received'] or 0),
            float((item['revenue'] or 0) - (item['received'] or 0)),
        ])
    
    exporter.add_sheet(SheetConfig(
        name="Revenue by Status",
        headers=["Status", "Orders", "Revenue", "Received", "Outstanding"],
        data=status_data
    ))
    
    # Sheet 3: Revenue by Branch
    branch_revenue = orders.values('branch__id', 'branch__name', 'branch__center__name').annotate(
        count=Count('id'),
        revenue=Coalesce(Sum('total_price'), Decimal('0'), output_field=DecimalField()),
        received=Coalesce(Sum('received'), Decimal('0'), output_field=DecimalField()),
        avg=Coalesce(Avg('total_price'), Decimal('0'), output_field=DecimalField())
    ).order_by('-revenue')
    
    branch_data = []
    for item in branch_revenue:
        if item['branch__id']:
            branch_data.append([
                item['branch__name'] or "Unassigned",
                item['branch__center__name'] or "N/A",
                item['count'],
                float(item['revenue'] or 0),
                float(item['received'] or 0),
                float((item['revenue'] or 0) - (item['received'] or 0)),
                float(item['avg'] or 0),
            ])
    
    exporter.add_sheet(SheetConfig(
        name="Revenue by Branch",
        headers=["Branch", "Center", "Orders", "Revenue", "Received", "Outstanding", "Avg Order Value"],
        data=branch_data
    ))
    
    # Sheet 4: Revenue by Product
    product_revenue = orders.values('product__name', 'product__category__name').annotate(
        count=Count('id'),
        revenue=Coalesce(Sum('total_price'), Decimal('0'), output_field=DecimalField()),
        pages=Coalesce(Sum('total_pages'), 0)
    ).order_by('-revenue')
    
    product_data = []
    for item in product_revenue:
        if item['product__name']:
            product_data.append([
                item['product__name'] or "Unknown",
                item['product__category__name'] or "N/A",
                item['count'],
                item['pages'] or 0,
                float(item['revenue'] or 0),
            ])
    
    exporter.add_sheet(SheetConfig(
        name="Revenue by Product",
        headers=["Product", "Category", "Orders", "Total Pages", "Revenue"],
        data=product_data
    ))
    
    # Sheet 5: Payment Method Breakdown
    payment_breakdown = orders.values('payment_type').annotate(
        count=Count('id'),
        revenue=Coalesce(Sum('total_price'), Decimal('0'), output_field=DecimalField()),
        received=Coalesce(Sum('received'), Decimal('0'), output_field=DecimalField())
    ).order_by('-revenue')
    
    payment_data = []
    for item in payment_breakdown:
        payment_type = item['payment_type'] or 'not_specified'
        pct = (item['revenue'] / total_revenue * 100) if total_revenue > 0 else 0
        collection = (item['received'] / item['revenue'] * 100) if item['revenue'] > 0 else 0
        payment_data.append([
            PAYMENT_TYPE_LABELS.get(payment_type, payment_type.title()),
            item['count'],
            f"{pct:.1f}%",
            float(item['revenue'] or 0),
            float(item['received'] or 0),
            float((item['revenue'] or 0) - (item['received'] or 0)),
            f"{collection:.1f}%",
        ])
    
    exporter.add_sheet(SheetConfig(
        name="Payment Methods",
        headers=["Payment Type", "Orders", "% of Total", "Revenue", "Received", "Outstanding", "Collection Rate"],
        data=payment_data,
        column_widths=[20, 10, 12, 15, 15, 15, 18]
    ))
    
    # Sheet 6: Daily Revenue Trend
    daily_revenue = orders.annotate(
        date=TruncDate('created_at')
    ).values('date').annotate(
        count=Count('id'),
        revenue=Coalesce(Sum('total_price'), Decimal('0'), output_field=DecimalField()),
        received=Coalesce(Sum('received'), Decimal('0'), output_field=DecimalField()),
        avg=Coalesce(Avg('total_price'), Decimal('0'), output_field=DecimalField())
    ).order_by('date')
    
    daily_data = []
    for item in daily_revenue:
        daily_data.append([
            item['date'],
            item['count'],
            float(item['revenue'] or 0),
            float(item['received'] or 0),
            float(item['avg'] or 0),
        ])
    
    exporter.add_sheet(SheetConfig(
        name="Daily Revenue Trend",
        headers=["Date", "Orders", "Revenue", "Received", "Avg Order Value"],
        data=daily_data
    ))
    
    # Sheet 6: Payment Details
    payments_data = []
    for order in orders.filter(received__gt=0).order_by('-created_at')[:500]:
        payments_data.append([
            order.id,
            order.created_at,
            order.bot_user.name if order.bot_user else "N/A",
            order.branch.name if order.branch else "N/A",
            float(order.total_price or 0),
            float(order.received or 0),
            float((order.total_price or 0) - (order.received or 0)),
            PAYMENT_TYPE_LABELS.get(order.payment_type, "N/A"),
            STATUS_LABELS.get(order.status, order.status),
        ])
    
    exporter.add_sheet(SheetConfig(
        name="Payment Details",
        headers=["Order ID", "Date", "Customer", "Branch", "Total", "Paid", "Outstanding", "Payment Type", "Status"],
        data=payments_data
    ))
    
    # Generate filename
    center_name = ""
    if center_id:
        try:
            center_name = f"_{TranslationCenter.objects.get(id=center_id).name}"
        except:
            pass
    
    filename = f"financial_report{center_name}_{date_from.strftime('%Y%m%d')}_to_{date_to.strftime('%Y%m%d')}.xlsx"
    
    return exporter.generate_response(filename)


def export_staff_performance(
    user,
    date_from: datetime,
    date_to: datetime,
    branch_id: Optional[str] = None,
    center_id: Optional[str] = None,
) -> HttpResponse:
    """
    Export Staff Performance Report with multiple sheets:
    - Executive Summary
    - Staff Summary
    - Performance Rankings
    - Detailed Orders by Staff
    - Daily Staff Activity
    """
    exporter = ExcelExporter()
    
    # Get filtered orders
    orders = get_user_orders(user).filter(
        created_at__gte=date_from,
        created_at__lte=date_to
    ).select_related('assigned_to__user', 'assigned_to__branch', 'bot_user', 'product')
    
    if center_id:
        orders = orders.filter(branch__center_id=center_id)
    if branch_id:
        orders = orders.filter(branch_id=branch_id)
    
    # Get staff members
    if user.is_superuser:
        staff_members = AdminUser.objects.filter(is_active=True)
        if center_id:
            staff_members = staff_members.filter(branch__center_id=center_id)
        if branch_id:
            staff_members = staff_members.filter(branch_id=branch_id)
    elif hasattr(user, 'admin_profile') and user.admin_profile:
        if user.admin_profile.is_owner:
            staff_members = AdminUser.objects.filter(
                center=user.admin_profile.center, is_active=True
            )
        elif user.admin_profile.is_manager:
            staff_members = AdminUser.objects.filter(
                branch=user.admin_profile.branch, is_active=True
            )
        else:
            staff_members = AdminUser.objects.filter(pk=user.admin_profile.pk)
    else:
        staff_members = AdminUser.objects.none()
    
    # Calculate overall statistics
    total_staff = staff_members.count()
    total_orders = orders.count()
    completed_orders = orders.filter(status='completed').count()
    total_revenue = orders.filter(status='completed').aggregate(total=Sum('total_price'))['total'] or 0
    total_pages = orders.aggregate(total=Sum('total_pages'))['total'] or 0
    avg_completion = (completed_orders / total_orders * 100) if total_orders > 0 else 0
    
    # Sheet 1: Executive Summary
    summary_data = [
        ["REPORT INFORMATION", ""],
        ["Report Type", "Staff Performance Report"],
        ["Generated Date", datetime.now().strftime('%Y-%m-%d %H:%M')],
        ["Period From", date_from.strftime('%Y-%m-%d')],
        ["Period To", date_to.strftime('%Y-%m-%d')],
        ["Generated By", user.get_full_name() or user.username],
        ["", ""],
        ["STAFF OVERVIEW", ""],
        ["Total Active Staff", total_staff],
        ["Staff with Assignments", orders.values('assigned_to').distinct().count()],
        ["Average Orders per Staff", round(total_orders / total_staff, 1) if total_staff > 0 else 0],
        ["", ""],
        ["PERFORMANCE METRICS", ""],
        ["Total Orders Assigned", total_orders],
        ["Orders Completed", completed_orders],
        ["Overall Completion Rate", f"{avg_completion:.1f}%"],
        ["Total Pages Processed", total_pages or 0],
        ["Total Revenue Generated", float(total_revenue)],
        ["Average Revenue per Staff", float(total_revenue / total_staff) if total_staff > 0 else 0],
        ["Average Pages per Staff", round(total_pages / total_staff, 1) if total_staff > 0 else 0],
    ]
    
    exporter.add_sheet(SheetConfig(
        name="Executive Summary",
        headers=["Metric", "Value"],
        data=summary_data,
        column_widths=[35, 25]
    ))
    
    # Sheet 2: Staff Summary
    staff_data = []
    for staff in staff_members:
        staff_orders = orders.filter(assigned_to=staff)
        total_assigned = staff_orders.count()
        completed = staff_orders.filter(status='completed').count()
        revenue = staff_orders.filter(status='completed').aggregate(total=Sum('total_price'))['total'] or 0
        pages = staff_orders.aggregate(total=Sum('total_pages'))['total'] or 0
        completion_rate = round(completed / total_assigned * 100, 1) if total_assigned > 0 else 0
        
        staff_data.append([
            staff.user.get_full_name() or staff.user.username,
            staff.branch.name if staff.branch else "N/A",
            staff.branch.center.name if staff.branch and staff.branch.center else "N/A",
            staff.role.name if staff.role else "Staff",
            total_assigned,
            completed,
            completion_rate,
            pages or 0,
            float(revenue),
        ])
    
    # Sort by completed orders
    staff_data.sort(key=lambda x: x[5], reverse=True)
    
    exporter.add_sheet(SheetConfig(
        name="Staff Summary",
        headers=["Staff Name", "Branch", "Center", "Role", "Total Assigned", "Completed", "Completion Rate (%)", "Total Pages", "Revenue"],
        data=staff_data
    ))
    
    # Sheet 3: Performance Rankings
    ranking_data = []
    for idx, staff_row in enumerate(staff_data, 1):
        ranking_data.append([
            idx,
            staff_row[0],  # Staff name
            staff_row[1],  # Branch
            staff_row[5],  # Completed
            staff_row[6],  # Completion rate
            staff_row[8],  # Revenue
            staff_row[7],  # Pages
            "🏆 Top Performer" if idx <= 3 else "⭐ High Performer" if idx <= 10 else "Good"
        ])
    
    exporter.add_sheet(SheetConfig(
        name="Performance Rankings",
        headers=["Rank", "Staff Name", "Branch", "Completed Orders", "Completion Rate (%)", "Revenue", "Pages", "Performance Level"],
        data=ranking_data,
        column_widths=[8, 25, 20, 18, 20, 15, 12, 20]
    ))
    
    # Sheet 4: Detailed Orders by Staff
    staff_orders_data = []
    for order in orders.filter(assigned_to__isnull=False).order_by('-created_at')[:1000]:
        staff_orders_data.append([
            order.id,
            order.created_at,
            order.assigned_to.user.get_full_name() if order.assigned_to else "N/A",
            order.bot_user.name if order.bot_user else "N/A",
            order.product.name if order.product else "N/A",
            order.total_pages or 0,
            float(order.total_price or 0),
            STATUS_LABELS.get(order.status, order.status),
            order.completed_at,
        ])
    
    exporter.add_sheet(SheetConfig(
        name="Orders by Staff",
        headers=["Order ID", "Created At", "Staff", "Customer", "Product", "Pages", "Price", "Status", "Completed At"],
        data=staff_orders_data
    ))
    
    # Sheet 5: Daily Staff Activity
    daily_staff = orders.filter(assigned_to__isnull=False).annotate(
        date=TruncDate('created_at')
    ).values('date', 'assigned_to__user__first_name', 'assigned_to__user__last_name').annotate(
        count=Count('id'),
        completed=Count('id', filter=Q(status='completed')),
        pages=Coalesce(Sum('total_pages'), 0),
        revenue=Coalesce(Sum('total_price'), Decimal('0'), output_field=DecimalField())
    ).order_by('date')
    
    daily_data = []
    for item in daily_staff:
        name = f"{item['assigned_to__user__first_name'] or ''} {item['assigned_to__user__last_name'] or ''}".strip() or "Unknown"
        daily_data.append([
            item['date'],
            name,
            item['count'],
            item['completed'],
            item['pages'] or 0,
            float(item['revenue'] or 0),
        ])
    
    exporter.add_sheet(SheetConfig(
        name="Daily Staff Activity",
        headers=["Date", "Staff", "Orders Assigned", "Orders Completed", "Total Pages", "Revenue"],
        data=daily_data
    ))
    
    # Generate filename
    filename = f"staff_performance_{date_from.strftime('%Y%m%d')}_to_{date_to.strftime('%Y%m%d')}.xlsx"
    
    return exporter.generate_response(filename)


def export_branch_comparison(
    user,
    date_from: datetime,
    date_to: datetime,
    center_id: Optional[str] = None,
) -> HttpResponse:
    """
    Export Branch Comparison Report with multiple sheets:
    - Executive Summary
    - Branch Summary
    - Market Share Analysis
    - Branch Daily Performance
    - Branch Staff Details
    - Branch Customer Distribution
    """
    exporter = ExcelExporter()
    
    # Get branches and orders
    branches = get_user_branches(user)
    orders = get_user_orders(user).filter(
        created_at__gte=date_from,
        created_at__lte=date_to
    )
    
    if center_id:
        branches = branches.filter(center_id=center_id)
        orders = orders.filter(branch__center_id=center_id)
    
    # Calculate overall statistics
    total_branches = branches.count()
    total_orders = orders.count()
    total_revenue = orders.aggregate(total=Sum('total_price'))['total'] or 0
    total_received = orders.aggregate(total=Sum('received'))['total'] or 0
    avg_branch_orders = total_orders / total_branches if total_branches > 0 else 0
    avg_branch_revenue = total_revenue / total_branches if total_branches > 0 else 0
    
    # Sheet 1: Executive Summary
    summary_data = [
        ["REPORT INFORMATION", ""],
        ["Report Type", "Branch Comparison Report"],
        ["Generated Date", datetime.now().strftime('%Y-%m-%d %H:%M')],
        ["Period From", date_from.strftime('%Y-%m-%d')],
        ["Period To", date_to.strftime('%Y-%m-%d')],
        ["Generated By", user.get_full_name() or user.username],
        ["", ""],
        ["BRANCH OVERVIEW", ""],
        ["Total Branches Analyzed", total_branches],
        ["Branches with Orders", orders.values('branch').distinct().count()],
        ["Average Orders per Branch", round(avg_branch_orders, 1)],
        ["Average Revenue per Branch", float(avg_branch_revenue)],
        ["", ""],
        ["OVERALL PERFORMANCE", ""],
        ["Total Orders", total_orders],
        ["Total Revenue", float(total_revenue)],
        ["Total Received", float(total_received)],
        ["Outstanding Balance", float(total_revenue - total_received)],
        ["Overall Collection Rate", f"{(total_received/total_revenue*100):.1f}%" if total_revenue > 0 else "0%"],
    ]
    
    exporter.add_sheet(SheetConfig(
        name="Executive Summary",
        headers=["Metric", "Value"],
        data=summary_data,
        column_widths=[35, 25]
    ))
    
    # Sheet 2: Branch Summary
    branch_data = []
    for branch in branches:
        branch_orders = orders.filter(branch=branch)
        total_orders = branch_orders.count()
        completed = branch_orders.filter(status='completed').count()
        revenue = branch_orders.aggregate(total=Sum('total_price'))['total'] or 0
        received = branch_orders.aggregate(total=Sum('received'))['total'] or 0
        avg_value = branch_orders.aggregate(avg=Avg('total_price'))['avg'] or 0
        staff_count = AdminUser.objects.filter(branch=branch, is_active=True).count()
        customer_count = BotUser.objects.filter(branch=branch, is_active=True).count()
        completion_rate = round(completed / total_orders * 100, 1) if total_orders > 0 else 0
        collection_rate = round(float(received) / float(revenue) * 100, 1) if revenue > 0 else 0
        
        branch_data.append([
            branch.name,
            branch.center.name if branch.center else "N/A",
            total_orders,
            completed,
            completion_rate,
            float(revenue),
            float(received),
            float(revenue - received),
            collection_rate,
            float(avg_value),
            staff_count,
            customer_count,
        ])
    
    # Sort by revenue
    branch_data.sort(key=lambda x: x[5], reverse=True)
    
    exporter.add_sheet(SheetConfig(
        name="Branch Summary",
        headers=[
            "Branch", "Center", "Total Orders", "Completed", "Completion Rate (%)",
            "Revenue", "Received", "Outstanding", "Collection Rate (%)",
            "Avg Order Value", "Staff Count", "Customer Count"
        ],
        data=branch_data
    ))
    
    # Sheet 3: Market Share Analysis
    market_share_data = []
    for branch_row in branch_data:
        revenue_share = (branch_row[5] / float(total_revenue) * 100) if total_revenue > 0 else 0
        order_share = (branch_row[2] / total_orders * 100) if total_orders > 0 else 0
        customer_share = (branch_row[11] / sum(b[11] for b in branch_data) * 100) if sum(b[11] for b in branch_data) > 0 else 0
        
        market_share_data.append([
            branch_row[0],  # Branch name
            branch_row[1],  # Center
            f"{revenue_share:.1f}%",
            f"{order_share:.1f}%",
            f"{customer_share:.1f}%",
            branch_row[5],  # Revenue
            branch_row[2],  # Orders
            branch_row[11],  # Customers
        ])
    
    exporter.add_sheet(SheetConfig(
        name="Market Share Analysis",
        headers=["Branch", "Center", "Revenue Share", "Order Share", "Customer Share", "Total Revenue", "Total Orders", "Total Customers"],
        data=market_share_data,
        column_widths=[20, 20, 15, 15, 18, 15, 15, 18]
    ))
    
    # Sheet 4: Branch Daily Performance
    daily_branch = orders.annotate(
        date=TruncDate('created_at')
    ).values('date', 'branch__name').annotate(
        count=Count('id'),
        revenue=Coalesce(Sum('total_price'), Decimal('0'), output_field=DecimalField()),
        received=Coalesce(Sum('received'), Decimal('0'), output_field=DecimalField())
    ).order_by('date', 'branch__name')
    
    daily_data = []
    for item in daily_branch:
        daily_data.append([
            item['date'],
            item['branch__name'] or "Unassigned",
            item['count'],
            float(item['revenue'] or 0),
            float(item['received'] or 0),
        ])
    
    exporter.add_sheet(SheetConfig(
        name="Daily Branch Performance",
        headers=["Date", "Branch", "Orders", "Revenue", "Received"],
        data=daily_data
    ))
    
    # Sheet 5: Branch Staff Details
    staff_data = []
    for branch in branches:
        staff_members = AdminUser.objects.filter(branch=branch, is_active=True)
        for staff in staff_members:
            staff_orders = orders.filter(assigned_to=staff)
            completed = staff_orders.filter(status='completed').count()
            revenue = staff_orders.aggregate(total=Sum('total_price'))['total'] or 0
            
            staff_data.append([
                branch.name,
                staff.user.get_full_name() or staff.user.username,
                staff.role.name if staff.role else "Staff",
                staff_orders.count(),
                completed,
                float(revenue),
            ])
    
    exporter.add_sheet(SheetConfig(
        name="Branch Staff Details",
        headers=["Branch", "Staff Name", "Role", "Orders Assigned", "Completed", "Revenue"],
        data=staff_data
    ))
    
    # Generate filename
    filename = f"branch_comparison_{date_from.strftime('%Y%m%d')}_to_{date_to.strftime('%Y%m%d')}.xlsx"
    
    return exporter.generate_response(filename)


def export_customer_analytics(
    user,
    date_from: datetime,
    date_to: datetime,
    branch_id: Optional[str] = None,
    center_id: Optional[str] = None,
) -> HttpResponse:
    """
    Export Customer Analytics Report with multiple sheets:
    - Executive Summary
    - Customer Segmentation
    - Top Customers
    - Customer List (Detailed)
    - B2B vs B2C Breakdown
    - New Customers
    - Customer Retention Analysis
    """
    exporter = ExcelExporter()
    
    # Get data based on user role
    customers = get_user_customers(user)
    orders = get_user_orders(user).filter(
        created_at__gte=date_from,
        created_at__lte=date_to
    )
    
    if center_id:
        customer_ids = orders.filter(branch__center_id=center_id).values_list('bot_user_id', flat=True).distinct()
        customers = customers.filter(id__in=customer_ids)
        orders = orders.filter(branch__center_id=center_id)
    if branch_id:
        customers = customers.filter(branch_id=branch_id)
        orders = orders.filter(branch_id=branch_id)
    
    # Calculate overall statistics
    total_customers = customers.count()
    active_customers = customers.filter(is_active=True).count()
    agencies = customers.filter(is_agency=True).count()
    new_customers = customers.filter(created_at__gte=date_from, created_at__lte=date_to).count()
    total_revenue = orders.aggregate(total=Sum('total_price'))['total'] or 0
    avg_customer_value = total_revenue / total_customers if total_customers > 0 else 0
    customers_with_orders = orders.values('bot_user').distinct().count()
    
    # Sheet 1: Executive Summary
    summary_data = [
        ["REPORT INFORMATION", ""],
        ["Report Type", "Customer Analytics Report"],
        ["Generated Date", datetime.now().strftime('%Y-%m-%d %H:%M')],
        ["Period From", date_from.strftime('%Y-%m-%d')],
        ["Period To", date_to.strftime('%Y-%m-%d')],
        ["Generated By", user.get_full_name() or user.username],
        ["", ""],
        ["CUSTOMER OVERVIEW", ""],
        ["Total Customers", total_customers],
        ["Active Customers", active_customers],
        ["Active Rate", f"{(active_customers/total_customers*100):.1f}%" if total_customers > 0 else "0%"],
        ["Customers with Orders (Period)", customers_with_orders],
        ["Customer Engagement Rate", f"{(customers_with_orders/total_customers*100):.1f}%" if total_customers > 0 else "0%"],
        ["New Customers (Period)", new_customers],
        ["", ""],
        ["CUSTOMER TYPES", ""],
        ["B2B (Agencies)", agencies],
        ["B2B Percentage", f"{(agencies/total_customers*100):.1f}%" if total_customers > 0 else "0%"],
        ["B2C (Regular)", total_customers - agencies],
        ["B2C Percentage", f"{((total_customers-agencies)/total_customers*100):.1f}%" if total_customers > 0 else "0%"],
        ["", ""],
        ["REVENUE METRICS", ""],
        ["Total Revenue (Period)", float(total_revenue)],
        ["Average Customer Value", float(avg_customer_value)],
        ["Average Orders per Customer", round(orders.count() / customers_with_orders, 1) if customers_with_orders > 0 else 0],
    ]
    
    exporter.add_sheet(SheetConfig(
        name="Executive Summary",
        headers=["Metric", "Value"],
        data=summary_data,
        column_widths=[35, 25]
    ))
    
    # Sheet 2: Customer Segmentation
    segmentation_data = []
    
    # High value customers (top 20% by revenue)
    high_value_threshold = total_revenue * 0.2 / customers_with_orders if customers_with_orders > 0 else 0
    high_value = orders.filter(bot_user__isnull=False).values('bot_user').annotate(
        total=Sum('total_price')
    ).filter(total__gte=high_value_threshold).count()
    
    segmentation_data.append([
        "High Value (Top 20%)",
        high_value,
        f"{(high_value/customers_with_orders*100):.1f}%" if customers_with_orders > 0 else "0%",
        f"≥ {float(high_value_threshold):.2f}",
        "VIP Treatment"
    ])
    
    # Medium value customers
    medium_value = customers_with_orders - high_value
    segmentation_data.append([
        "Medium Value",
        medium_value,
        f"{(medium_value/customers_with_orders*100):.1f}%" if customers_with_orders > 0 else "0%",
        f"< {float(high_value_threshold):.2f}",
        "Standard Service"
    ])
    
    # Inactive customers
    inactive = total_customers - customers_with_orders
    segmentation_data.append([
        "Inactive (No Orders)",
        inactive,
        f"{(inactive/total_customers*100):.1f}%" if total_customers > 0 else "0%",
        "0",
        "Re-engagement Needed"
    ])
    
    exporter.add_sheet(SheetConfig(
        name="Customer Segmentation",
        headers=["Segment", "Customer Count", "% of Total", "Revenue Threshold", "Recommended Action"],
        data=segmentation_data,
        column_widths=[25, 18, 15, 20, 25]
    ))
    
    # Sheet 3: Top Customers by Revenue
    top_customers = orders.filter(bot_user__isnull=False).values(
        'bot_user__id', 'bot_user__name', 'bot_user__phone', 'bot_user__is_agency'
    ).annotate(
        order_count=Count('id'),
        total_spent=Coalesce(Sum('total_price'), Decimal('0'), output_field=DecimalField()),
        total_paid=Coalesce(Sum('received'), Decimal('0'), output_field=DecimalField())
    ).order_by('-total_spent')[:50]
    
    top_data = []
    for item in top_customers:
        top_data.append([
            item['bot_user__name'] or "Unknown",
            item['bot_user__phone'] or "N/A",
            "B2B" if item['bot_user__is_agency'] else "B2C",
            item['order_count'],
            float(item['total_spent'] or 0),
            float(item['total_paid'] or 0),
            float((item['total_spent'] or 0) - (item['total_paid'] or 0)),
        ])
    
    exporter.add_sheet(SheetConfig(
        name="Top Customers",
        headers=["Customer Name", "Phone", "Type", "Orders", "Total Spent", "Paid", "Outstanding"],
        data=top_data
    ))
    
    # Sheet 4: Detailed Customer List
    customer_data = []
    for customer in customers.select_related('branch', 'agency')[:500]:
        customer_orders = orders.filter(bot_user=customer)
        order_count = customer_orders.count()
        total_spent = customer_orders.aggregate(total=Sum('total_price'))['total'] or 0
        
        customer_data.append([
            customer.id,
            customer.name,
            customer.phone,
            customer.username or "N/A",
            "B2B" if customer.is_agency else "B2C",
            customer.branch.name if customer.branch else "N/A",
            customer.agency.name if customer.agency else "N/A",
            customer.language,
            order_count,
            float(total_spent),
            customer.created_at,
            "Active" if customer.is_active else "Inactive",
        ])
    
    exporter.add_sheet(SheetConfig(
        name="Customer List",
        headers=[
            "ID", "Name", "Phone", "Username", "Type", "Branch", "Agency",
            "Language", "Total Orders", "Total Spent", "Registered At", "Status"
        ],
        data=customer_data
    ))
    
    # Sheet 5: B2B vs B2C Breakdown
    b2b_orders = orders.filter(bot_user__isnull=False, bot_user__is_agency=True)
    b2c_orders = orders.filter(bot_user__isnull=False, bot_user__is_agency=False)
    
    type_data = [
        [
            "B2B (Agencies)",
            b2b_orders.count(),
            float(b2b_orders.aggregate(total=Sum('total_price'))['total'] or 0),
            float(b2b_orders.aggregate(total=Sum('received'))['total'] or 0),
            agencies,
        ],
        [
            "B2C (Regular)",
            b2c_orders.count(),
            float(b2c_orders.aggregate(total=Sum('total_price'))['total'] or 0),
            float(b2c_orders.aggregate(total=Sum('received'))['total'] or 0),
            total_customers - agencies,
        ],
    ]
    
    exporter.add_sheet(SheetConfig(
        name="B2B vs B2C",
        headers=["Customer Type", "Orders", "Revenue", "Paid", "Customer Count"],
        data=type_data
    ))
    
    # Sheet 5: New Customers (Period)
    new_customer_list = customers.filter(
        created_at__gte=date_from, created_at__lte=date_to
    ).select_related('branch')
    
    new_data = []
    for customer in new_customer_list:
        new_data.append([
            customer.name,
            customer.phone,
            "B2B" if customer.is_agency else "B2C",
            customer.branch.name if customer.branch else "N/A",
            customer.created_at,
        ])
    
    exporter.add_sheet(SheetConfig(
        name="New Customers",
        headers=["Name", "Phone", "Type", "Branch", "Registered At"],
        data=new_data
    ))
    
    # Sheet 7: Customer Retention Analysis
    retention_data = []
    
    # Repeat customers (more than 1 order in period)
    repeat_customers = orders.values('bot_user').annotate(
        order_count=Count('id')
    ).filter(order_count__gt=1).count()
    
    # One-time customers
    onetime_customers = orders.values('bot_user').annotate(
        order_count=Count('id')
    ).filter(order_count=1).count()
    
    retention_data.append([
        "Repeat Customers",
        repeat_customers,
        f"{(repeat_customers/customers_with_orders*100):.1f}%" if customers_with_orders > 0 else "0%",
        "Loyal customer base"
    ])
    
    retention_data.append([
        "One-Time Customers",
        onetime_customers,
        f"{(onetime_customers/customers_with_orders*100):.1f}%" if customers_with_orders > 0 else "0%",
        "Retention opportunity"
    ])
    
    retention_data.append([
        "Registered but Inactive",
        total_customers - customers_with_orders,
        f"{((total_customers-customers_with_orders)/total_customers*100):.1f}%" if total_customers > 0 else "0%",
        "Need activation campaign"
    ])
    
    exporter.add_sheet(SheetConfig(
        name="Retention Analysis",
        headers=["Customer Type", "Count", "Percentage", "Notes"],
        data=retention_data,
        column_widths=[25, 12, 15, 30]
    ))

    # Generate filename
    filename = f"customer_analytics_{date_from.strftime('%Y%m%d')}_to_{date_to.strftime('%Y%m%d')}.xlsx"
    
    return exporter.generate_response(filename)


def export_unit_economy(
    user,
    branch_id: Optional[str] = None,
    center_id: Optional[str] = None,
) -> HttpResponse:
    """
    Export Unit Economy Report with multiple sheets:
    - Executive Summary
    - Remaining Balance Summary
    - Remaining by Branch
    - Remaining by Client Type
    - Remaining by Center
    - Top Debtors
    - Payment Behavior Analysis
    - Detailed Outstanding Orders
    """
    from services.analytics import (
        get_remaining_balance_summary,
        get_remaining_by_branch,
        get_remaining_by_client_type,
        get_remaining_by_center,
        get_top_debtors,
    )
    from decimal import Decimal
    from django.db.models.functions import Coalesce
    from django.db.models import Case, When, F, DecimalField
    
    exporter = ExcelExporter()
    
    # Get analytics data
    summary = get_remaining_balance_summary(user)
    by_branch = get_remaining_by_branch(user, limit=100)
    by_client_type = get_remaining_by_client_type(user)
    by_center = get_remaining_by_center(user, limit=50)
    top_debtors = get_top_debtors(user, limit=100)
    
    # Sheet 1: Executive Summary
    summary_data = [
        ["REPORT INFORMATION", ""],
        ["Report Type", "Unit Economy Report"],
        ["Generated Date", datetime.now().strftime('%Y-%m-%d %H:%M')],
        ["Generated By", user.get_full_name() or user.username],
        ["", ""],
        ["BALANCE OVERVIEW", ""],
        ["Total Outstanding Balance", summary['total_remaining']],
        ["Orders with Outstanding Balance", summary['total_orders_with_debt']],
        ["Fully Paid Orders", summary['fully_paid_count']],
        ["Debt Ratio", f"{(summary['total_orders_with_debt']/(summary['total_orders_with_debt']+summary['fully_paid_count'])*100):.1f}%" if (summary['total_orders_with_debt']+summary['fully_paid_count']) > 0 else "0%"],
        ["", ""],
        ["REVENUE METRICS", ""],
        ["Total Received", summary['total_received']],
        ["Total Expected Revenue", summary['total_expected']],
        ["Collection Rate", f"{summary['collection_rate']:.1f}%"],
        ["Revenue at Risk", summary['total_remaining']],
        ["Average Outstanding per Order", summary['total_remaining'] / summary['total_orders_with_debt'] if summary['total_orders_with_debt'] > 0 else 0],
        ["", ""],
        ["RECOMMENDATIONS", ""],
        ["Payment Collection Priority", "High" if summary['collection_rate'] < 70 else "Medium" if summary['collection_rate'] < 85 else "Low"],
        ["Risk Level", "High Risk" if summary['collection_rate'] < 60 else "Moderate" if summary['collection_rate'] < 80 else "Low Risk"],
    ]
    
    exporter.add_sheet(SheetConfig(
        name="Executive Summary",
        headers=["Metric", "Value"],
        data=summary_data,
        column_widths=[35, 25]
    ))
    
    # Sheet 2: Summary
    summary_data = [
        ["Total Outstanding Balance", summary['total_remaining']],
        ["Orders with Outstanding Balance", summary['total_orders_with_debt']],
        ["Fully Paid Orders", summary['fully_paid_count']],
        ["Total Received", summary['total_received']],
        ["Total Expected Revenue", summary['total_expected']],
        ["Collection Rate (%)", summary['collection_rate']],
    ]
    
    exporter.add_sheet(SheetConfig(
        name="Summary",
        headers=["Metric", "Value"],
        data=summary_data,
        column_widths=[35, 20]
    ))
    
    # Sheet 2: By Branch
    branch_data = []
    for item in by_branch:
        branch_data.append([
            item['branch_name'],
            item['center_name'],
            item['total_orders'],
            item['orders_with_debt'],
            item['remaining'],
            item['total_received'],
            item['total_expected'],
            item['collection_rate'],
        ])
    
    exporter.add_sheet(SheetConfig(
        name="By Branch",
        headers=["Branch", "Center", "Total Orders", "Orders with Debt", "Outstanding", "Received", "Expected", "Collection Rate (%)"],
        data=branch_data
    ))
    
    # Sheet 3: By Client Type
    type_data = [
        [
            by_client_type['agency']['label'],
            by_client_type['agency']['total_orders'],
            by_client_type['agency']['orders_with_debt'],
            by_client_type['agency']['remaining'],
            by_client_type['agency']['total_received'],
            by_client_type['agency']['collection_rate'],
        ],
        [
            by_client_type['regular']['label'],
            by_client_type['regular']['total_orders'],
            by_client_type['regular']['orders_with_debt'],
            by_client_type['regular']['remaining'],
            by_client_type['regular']['total_received'],
            by_client_type['regular']['collection_rate'],
        ],
    ]
    
    exporter.add_sheet(SheetConfig(
        name="By Client Type",
        headers=["Client Type", "Total Orders", "Orders with Debt", "Outstanding", "Received", "Collection Rate (%)"],
        data=type_data
    ))
    
    # Sheet 4: By Center
    center_data = []
    for item in by_center:
        center_data.append([
            item['center_name'],
            item['total_orders'],
            item['orders_with_debt'],
            item['remaining'],
            item['total_received'],
            item['collection_rate'],
        ])
    
    exporter.add_sheet(SheetConfig(
        name="By Center",
        headers=["Center", "Total Orders", "Orders with Debt", "Outstanding", "Received", "Collection Rate (%)"],
        data=center_data
    ))
    
    # Sheet 5: Top Debtors
    debtor_data = []
    for item in top_debtors:
        debtor_data.append([
            item['customer_name'],
            item['customer_phone'],
            "B2B" if item['is_agency'] else "B2C",
            item['total_orders'],
            item['orders_with_debt'],
            item['remaining'],
            item['total_expected'],
            item['total_received'],
        ])
    
    exporter.add_sheet(SheetConfig(
        name="Top Debtors",
        headers=["Customer Name", "Phone", "Type", "Total Orders", "Orders with Debt", "Outstanding", "Total Expected", "Total Paid"],
        data=debtor_data
    ))
    
    # Sheet 7: Payment Behavior Analysis
    orders_all = get_user_orders(user).exclude(status='cancelled')
    
    # Categorize payment behavior
    fully_paid = orders_all.filter(payment_accepted_fully=True).count()
    partial_paid = orders_all.filter(received__gt=0, payment_accepted_fully=False).count()
    unpaid = orders_all.filter(received=0).count()
    total_analyzed = fully_paid + partial_paid + unpaid
    
    behavior_data = [
        [
            "Fully Paid Orders",
            fully_paid,
            f"{(fully_paid/total_analyzed*100):.1f}%" if total_analyzed > 0 else "0%",
            "Excellent payment behavior",
            "No action required"
        ],
        [
            "Partially Paid Orders",
            partial_paid,
            f"{(partial_paid/total_analyzed*100):.1f}%" if total_analyzed > 0 else "0%",
            "Payment in progress",
            "Follow up on remaining balance"
        ],
        [
            "Unpaid Orders",
            unpaid,
            f"{(unpaid/total_analyzed*100):.1f}%" if total_analyzed > 0 else "0%",
            "No payment received",
            "Urgent: Contact customer immediately"
        ],
    ]
    
    exporter.add_sheet(SheetConfig(
        name="Payment Behavior",
        headers=["Payment Status", "Order Count", "Percentage", "Description", "Recommended Action"],
        data=behavior_data,
        column_widths=[25, 15, 15, 25, 35]
    ))
    
    # Sheet 8: Detailed Outstanding Orders
    orders = get_user_orders(user).exclude(status='cancelled').select_related(
        'bot_user', 'product', 'branch'
    ).annotate(
        calc_remaining=Case(
            When(payment_accepted_fully=True, then=Decimal('0')),
            default=(
                Coalesce(F('total_price'), Decimal('0')) + Coalesce(F('extra_fee'), Decimal('0'))
            ) - Coalesce(F('received'), Decimal('0')),
            output_field=DecimalField(max_digits=12, decimal_places=2)
        )
    ).filter(calc_remaining__gt=0).order_by('-calc_remaining')[:500]
    
    orders_data = []
    for order in orders:
        remaining = float(order.calc_remaining) if hasattr(order, 'calc_remaining') else 0
        orders_data.append([
            order.id,
            order.created_at,
            order.bot_user.name if order.bot_user else "N/A",
            order.bot_user.phone if order.bot_user else "N/A",
            "B2B" if (order.bot_user and order.bot_user.is_agency) else "B2C",
            order.branch.name if order.branch else "N/A",
            order.product.name if order.product else "N/A",
            float(order.total_price or 0),
            float(order.extra_fee or 0),
            float(order.received or 0),
            remaining,
            STATUS_LABELS.get(order.status, order.status),
        ])
    
    exporter.add_sheet(SheetConfig(
        name="Outstanding Orders",
        headers=["Order ID", "Date", "Customer", "Phone", "Type", "Branch", "Product", "Total Price", "Extra Fee", "Received", "Outstanding", "Status"],
        data=orders_data
    ))
    
    # Generate filename
    today = timezone.now().strftime('%Y%m%d')
    filename = f"unit_economy_{today}.xlsx"
    
    return exporter.generate_response(filename)


def export_my_statistics(
    user,
    date_from: datetime,
    date_to: datetime,
) -> HttpResponse:
    """
    Export My Statistics Report with multiple sheets:
    - Executive Summary
    - Performance Metrics
    - My Orders
    - Daily Performance
    - Weekly Trends
    """
    exporter = ExcelExporter()
    
    # Get admin profile
    admin_profile = getattr(user, 'admin_profile', None)
    
    if admin_profile:
        my_orders = Order.objects.filter(assigned_to=admin_profile).select_related(
            'bot_user', 'product', 'branch'
        )
    else:
        my_orders = Order.objects.none()
    
    # Filter by period
    period_orders = my_orders.filter(created_at__gte=date_from, created_at__lte=date_to)
    
    # Calculate statistics
    total_count = period_orders.count()
    completed = period_orders.filter(status='completed').count()
    pending = period_orders.filter(status='pending').count()
    in_progress = period_orders.filter(status='in_progress').count()
    total_pages = period_orders.aggregate(total=Sum('total_pages'))['total'] or 0
    total_revenue = period_orders.aggregate(total=Sum('total_price'))['total'] or 0
    completion_rate = round(completed / total_count * 100, 1) if total_count > 0 else 0
    avg_daily = total_count / ((date_to - date_from).days + 1)
    
    # Sheet 1: Executive Summary
    summary_data = [
        ["REPORT INFORMATION", ""],
        ["Report Type", "Personal Statistics Report"],
        ["Generated Date", datetime.now().strftime('%Y-%m-%d %H:%M')],
        ["Period From", date_from.strftime('%Y-%m-%d')],
        ["Period To", date_to.strftime('%Y-%m-%d')],
        ["Staff Member", user.get_full_name() or user.username],
        ["Branch", admin_profile.branch.name if (admin_profile and admin_profile.branch) else "N/A"],
        ["", ""],
        ["PERFORMANCE OVERVIEW", ""],
        ["Total Orders Assigned", total_count],
        ["Orders Completed", completed],
        ["Completion Rate", f"{completion_rate}%"],
        ["Orders Pending", pending],
        ["Orders In Progress", in_progress],
        ["Average Orders per Day", round(avg_daily, 1)],
        ["", ""],
        ["PRODUCTIVITY METRICS", ""],
        ["Total Pages Processed", total_pages],
        ["Average Pages per Order", round(total_pages / total_count, 1) if total_count > 0 else 0],
        ["Total Revenue Generated", float(total_revenue)],
        ["Average Revenue per Order", float(total_revenue / total_count) if total_count > 0 else 0],
        ["Revenue per Page", float(total_revenue / total_pages) if total_pages > 0 else 0],
        ["", ""],
        ["PERFORMANCE RATING", ""],
        ["Completion Rate Rating", "Excellent" if completion_rate >= 90 else "Good" if completion_rate >= 75 else "Needs Improvement"],
        ["Productivity Level", "High" if avg_daily >= 10 else "Medium" if avg_daily >= 5 else "Low"],
    ]
    
    exporter.add_sheet(SheetConfig(
        name="Executive Summary",
        headers=["Metric", "Value"],
        data=summary_data,
        column_widths=[35, 25]
    ))
    
    # Sheet 2: Performance Metrics by Status
    status_breakdown = period_orders.values('status').annotate(
        count=Count('id'),
        pages=Coalesce(Sum('total_pages'), 0),
        revenue=Coalesce(Sum('total_price'), 0)
    ).order_by('-count')
    
    status_data = []
    for item in status_breakdown:
        pct = (item['count'] / total_count * 100) if total_count > 0 else 0
        status_data.append([
            STATUS_LABELS.get(item['status'], item['status']),
            item['count'],
            f"{pct:.1f}%",
            item['pages'] or 0,
            float(item['revenue'] or 0),
        ])
    
    exporter.add_sheet(SheetConfig(
        name="Performance Metrics",
        headers=["Status", "Order Count", "% of Total", "Pages", "Revenue"],
        data=status_data,
        column_widths=[20, 15, 12, 12, 15]
    ))
    
    # Sheet 2: My Orders
    orders_data = []
    for order in period_orders.order_by('-created_at'):
        orders_data.append([
            order.id,
            order.created_at,
            order.bot_user.name if order.bot_user else "N/A",
            order.product.name if order.product else "N/A",
            order.total_pages or 0,
            float(order.total_price or 0),
            STATUS_LABELS.get(order.status, order.status),
            order.completed_at,
        ])
    
    exporter.add_sheet(SheetConfig(
        name="My Orders",
        headers=["Order ID", "Created At", "Customer", "Product", "Pages", "Price", "Status", "Completed At"],
        data=orders_data
    ))
    
    # Sheet 3: Daily Performance
    daily_performance = period_orders.annotate(
        date=TruncDate('created_at')
    ).values('date').annotate(
        count=Count('id'),
        completed=Count('id', filter=Q(status='completed')),
        pages=Coalesce(Sum('total_pages'), 0),
        revenue=Coalesce(Sum('total_price'), 0)
    ).order_by('date')
    
    daily_data = []
    for item in daily_performance:
        daily_data.append([
            item['date'],
            item['count'],
            item['completed'],
            item['pages'] or 0,
            float(item['revenue'] or 0),
        ])
    
    exporter.add_sheet(SheetConfig(
        name="Daily Performance",
        headers=["Date", "Orders Assigned", "Completed", "Pages", "Revenue"],
        data=daily_data
    ))
    
    # Sheet 5: Weekly Trends (if period > 7 days)
    if (date_to - date_from).days >= 7:
        from django.db.models.functions import ExtractWeek, ExtractYear
        
        weekly_performance = period_orders.annotate(
            year=ExtractYear('created_at'),
            week=ExtractWeek('created_at')
        ).values('year', 'week').annotate(
            count=Count('id'),
            completed=Count('id', filter=Q(status='completed')),
            pages=Coalesce(Sum('total_pages'), 0),
            revenue=Coalesce(Sum('total_price'), 0)
        ).order_by('year', 'week')
        
        weekly_data = []
        for item in weekly_performance:
            weekly_data.append([
                f"{item['year']}-W{item['week']:02d}",
                item['count'],
                item['completed'],
                f"{(item['completed']/item['count']*100):.1f}%" if item['count'] > 0 else "0%",
                item['pages'] or 0,
                float(item['revenue'] or 0),
            ])
        
        exporter.add_sheet(SheetConfig(
            name="Weekly Trends",
            headers=["Week", "Orders", "Completed", "Completion Rate", "Pages", "Revenue"],
            data=weekly_data,
            column_widths=[15, 10, 12, 18, 12, 15]
        ))
    
    # Generate filename
    staff_name = user.get_full_name() or user.username
    filename = f"my_statistics_{staff_name}_{date_from.strftime('%Y%m%d')}_to_{date_to.strftime('%Y%m%d')}.xlsx"
    filename = filename.replace(" ", "_")
    
    return exporter.generate_response(filename)


def export_expense_analytics(
    user,
    date_from: datetime,
    date_to: datetime,
    branch_id: Optional[str] = None,
    center_id: Optional[str] = None,
    expense_type_filter: Optional[str] = None,
) -> HttpResponse:
    """
    Export Expense Analytics Report with multiple sheets:
    - Expense Summary
    - By Branch
    - By Center (if superuser)
    - Top Expenses
    - Expense Type Distribution
    """
    from services.models import Expense
    from organizations.rbac import get_user_expenses
    from orders.models import Order
    
    exporter = ExcelExporter()
    
    # Get user access level
    user_profile = getattr(user, 'admin_profile', None)
    is_superuser = user.is_superuser
    is_center_owner = False
    
    if user_profile:
        is_center_owner = (user_profile.is_owner or 
                          (user_profile.center and user_profile.center.owner_id == user.id))
    
    # Get accessible branches
    accessible_branches = get_user_branches(user)
    
    # Filter orders by date range
    orders_base = Order.objects.filter(
        created_at__gte=date_from,
        created_at__lte=date_to
    ).select_related('product', 'branch', 'branch__center', 'bot_user')
    
    # Apply access filters
    if is_superuser:
        if center_id:
            orders_base = orders_base.filter(branch__center_id=center_id)
    elif is_center_owner and user_profile.center:
        orders_base = orders_base.filter(branch__center=user_profile.center)
    else:
        orders_base = orders_base.filter(branch__in=accessible_branches)
    
    if is_superuser and branch_id:
        orders_base = orders_base.filter(branch_id=branch_id)
    
    # Get expenses with access filtering
    expenses = get_user_expenses(user).select_related('branch', 'branch__center')
    
    # Apply same filters to expenses
    if is_superuser and center_id:
        expenses = expenses.filter(branch__center_id=center_id)
    elif is_center_owner and user_profile.center:
        expenses = expenses.filter(branch__center=user_profile.center)
    else:
        expenses = expenses.filter(branch__in=accessible_branches)
    
    if is_superuser and branch_id:
        expenses = expenses.filter(branch_id=branch_id)
    
    # Apply expense type filter
    if expense_type_filter in ['b2b', 'b2c']:
        expenses = expenses.filter(expense_type__in=[expense_type_filter, 'both'])
    
    # Calculate expense usage and totals
    expense_usage = {}
    total_expense_cost = Decimal('0')
    b2b_expense_cost = Decimal('0')
    b2c_expense_cost = Decimal('0')
    both_expense_cost = Decimal('0')
    
    for expense in expenses:
        products_with_expense = expense.products.all()
        related_orders = orders_base.filter(product__in=products_with_expense)
        order_count = related_orders.count()
        expense_total = Decimal('0')
        for order in related_orders:
            expense_total += expense.price_for_original + (expense.price_for_copy * (order.copy_number or 0))
        
        expense_usage[expense.id] = {
            'expense': expense,
            'order_count': order_count,
            'total_cost': expense_total
        }
        
        total_expense_cost += expense_total
        
        if expense.expense_type == 'b2b':
            b2b_expense_cost += expense_total
        elif expense.expense_type == 'b2c':
            b2c_expense_cost += expense_total
        else:
            both_expense_cost += expense_total
    
    # Calculate statistics for all expense items
    total_count = sum(e['order_count'] for e in expense_usage.values())
    b2b_count = sum(e['order_count'] for e in expense_usage.values() if e['expense'].expense_type == 'b2b')
    b2c_count = sum(e['order_count'] for e in expense_usage.values() if e['expense'].expense_type == 'b2c')
    both_count = sum(e['order_count'] for e in expense_usage.values() if e['expense'].expense_type == 'both')
    
    unique_expenses = len(expense_usage)
    active_expenses = len([e for e in expense_usage.values() if e['order_count'] > 0])
    
    # Calculate percentages
    b2b_percentage = (b2b_expense_cost / total_expense_cost * 100) if total_expense_cost > 0 else 0
    b2c_percentage = (b2c_expense_cost / total_expense_cost * 100) if total_expense_cost > 0 else 0
    both_percentage = (both_expense_cost / total_expense_cost * 100) if total_expense_cost > 0 else 0
    
    # Sheet 1: Executive Summary with Report Info
    summary_data = [
        ["REPORT INFORMATION", ""],
        ["Report Type", "Expense Analytics Report"],
        ["Generated Date", datetime.now().strftime('%Y-%m-%d %H:%M')],
        ["Period From", date_from.strftime('%Y-%m-%d')],
        ["Period To", date_to.strftime('%Y-%m-%d')],
        ["Generated By", user.get_full_name() or user.username],
        ["", ""],
        ["EXPENSE OVERVIEW", ""],
        ["Total Expense Cost", float(total_expense_cost)],
        ["Total Orders with Expenses", total_count],
        ["Unique Expense Types Used", unique_expenses],
        ["Active Expenses (with orders)", active_expenses],
        ["Average Cost per Order", float(total_expense_cost / total_count) if total_count > 0 else 0],
        ["", ""],
        ["B2B EXPENSES", ""],
        ["B2B Total Cost", float(b2b_expense_cost)],
        ["B2B Percentage of Total", f"{b2b_percentage:.1f}%"],
        ["B2B Order Count", b2b_count],
        ["B2B Avg per Order", float(b2b_expense_cost / b2b_count) if b2b_count > 0 else 0],
        ["", ""],
        ["B2C EXPENSES", ""],
        ["B2C Total Cost", float(b2c_expense_cost)],
        ["B2C Percentage of Total", f"{b2c_percentage:.1f}%"],
        ["B2C Order Count", b2c_count],
        ["B2C Avg per Order", float(b2c_expense_cost / b2c_count) if b2c_count > 0 else 0],
        ["", ""],
        ["BOTH TYPES EXPENSES", ""],
        ["Both Types Total Cost", float(both_expense_cost)],
        ["Both Types Percentage", f"{both_percentage:.1f}%"],
        ["Both Types Order Count", both_count],
        ["Both Types Avg per Order", float(both_expense_cost / both_count) if both_count > 0 else 0],
    ]
    
    exporter.add_sheet(SheetConfig(
        name="Executive Summary",
        headers=["Metric", "Value"],
        data=summary_data,
        column_widths=[35, 25]
    ))
    
    # Sheet 2: All Expenses Detail (Complete List)
    all_expenses_data = []
    for expense_id, data in sorted(expense_usage.items(), key=lambda x: x[1]['total_cost'], reverse=True):
        expense = data['expense']
        type_label = {
            'b2b': 'B2B (Agency/Business)',
            'b2c': 'B2C (Individual)',
            'both': 'Both B2B & B2C'
        }.get(expense.expense_type, 'N/A')
        
        # Get list of products using this expense
        products_list = ", ".join([p.name for p in expense.products.all()[:5]])
        if expense.products.count() > 5:
            products_list += f" (+{expense.products.count() - 5} more)"
        
        all_expenses_data.append([
            expense.name,
            expense.description or "",
            type_label,
            float(expense.price_for_original),
            float(expense.price_for_copy),
            data['order_count'],
            float(data['total_cost']),
            float(data['total_cost'] / data['order_count']) if data['order_count'] > 0 else 0,
            expense.branch.name if expense.branch else "N/A",
            expense.branch.center.name if (expense.branch and expense.branch.center) else "N/A",
            products_list or "No products assigned",
            "Active" if data['order_count'] > 0 else "Inactive",
        ])
    
    exporter.add_sheet(SheetConfig(
        name="All Expenses Detail",
        headers=[
            "Expense Name", "Description", "Type", "Price (Original)", "Price (Per Copy)",
            "Orders", "Total Cost", "Avg Cost/Order", 
            "Branch", "Center", "Products Using", "Status"
        ],
        data=all_expenses_data,
        column_widths=[25, 30, 22, 12, 10, 15, 15, 20, 20, 35, 12]
    ))
    
    # Sheet 3: By Branch (Enhanced)
    by_branch_data = []
    for branch in accessible_branches:
        branch_orders = orders_base.filter(branch=branch)
        branch_expenses = expenses.filter(branch=branch)
        
        branch_total = Decimal('0')
        branch_count = 0
        branch_b2b = Decimal('0')
        branch_b2c = Decimal('0')
        branch_both = Decimal('0')
        unique_branch_expenses = 0
        
        for expense in branch_expenses:
            products_with_expense = expense.products.all()
            related_orders = branch_orders.filter(product__in=products_with_expense)
            order_count = related_orders.count()
            expense_total = Decimal('0')
            for order in related_orders:
                expense_total += expense.price_for_original + (expense.price_for_copy * (order.copy_number or 0))
            
            if order_count > 0:
                unique_branch_expenses += 1
            
            branch_total += expense_total
            branch_count += order_count
            
            if expense.expense_type == 'b2b':
                branch_b2b += expense_total
            elif expense.expense_type == 'b2c':
                branch_b2c += expense_total
            else:
                branch_both += expense_total
        
        if branch_total > 0:
            branch_pct = (branch_total / total_expense_cost * 100) if total_expense_cost > 0 else 0
            by_branch_data.append([
                branch.name,
                branch.center.name if branch.center else "N/A",
                float(branch_total),
                f"{branch_pct:.1f}%",
                branch_count,
                unique_branch_expenses,
                float(branch_total / branch_count) if branch_count > 0 else 0,
                float(branch_b2b),
                float(branch_b2c),
                float(branch_both),
            ])
    
    by_branch_data.sort(key=lambda x: x[2], reverse=True)
    
    exporter.add_sheet(SheetConfig(
        name="By Branch Analysis",
        headers=[
            "Branch", "Center", "Total Cost", "% of Total", 
            "Orders", "Active Expenses", "Avg per Order", 
            "B2B Cost", "B2C Cost", "Both Cost"
        ],
        data=by_branch_data,
        column_widths=[20, 20, 15, 12, 10, 15, 15, 15, 15, 15]
    ))
    
    # Sheet 4: Top 30 Expenses by Cost
    top_expenses_data = []
    rank = 1
    for expense_id, data in sorted(expense_usage.items(), key=lambda x: x[1]['total_cost'], reverse=True)[:30]:
        expense = data['expense']
        type_label = {
            'b2b': 'B2B',
            'b2c': 'B2C',
            'both': 'Both'
        }.get(expense.expense_type, 'N/A')
        
        cost_pct = (data['total_cost'] / total_expense_cost * 100) if total_expense_cost > 0 else 0
        
        top_expenses_data.append([
            rank,
            expense.name,
            type_label,
            float(expense.price_for_original),
            float(expense.price_for_copy),
            data['order_count'],
            float(data['total_cost']),
            f"{cost_pct:.1f}%",
            float(data['total_cost'] / data['order_count']) if data['order_count'] > 0 else 0,
            expense.branch.name if expense.branch else "N/A",
            expense.description or "",
        ])
        rank += 1
    
    exporter.add_sheet(SheetConfig(
        name="Top 30 Expenses",
        headers=[
            "Rank", "Expense Name", "Type", "Price (Original)", "Price (Per Copy)",
            "Orders", "Total Cost", "% of Total", "Avg/Order", 
            "Branch", "Description"
        ],
        data=top_expenses_data,
        column_widths=[8, 25, 12, 12, 10, 15, 12, 12, 20, 30]
    ))
    
    # Sheet 5: Type Distribution with Details
    type_data = []
    if b2b_expense_cost > 0:
        b2b_avg = b2b_expense_cost / b2b_count if b2b_count > 0 else 0
        type_data.append([
            "B2B (Agency/Business)", 
            float(b2b_expense_cost), 
            f"{b2b_percentage:.1f}%",
            b2b_count,
            float(b2b_avg),
            len([e for e in expense_usage.values() if e['expense'].expense_type == 'b2b' and e['order_count'] > 0])
        ])
    if b2c_expense_cost > 0:
        b2c_avg = b2c_expense_cost / b2c_count if b2c_count > 0 else 0
        type_data.append([
            "B2C (Individual Customer)", 
            float(b2c_expense_cost), 
            f"{b2c_percentage:.1f}%",
            b2c_count,
            float(b2c_avg),
            len([e for e in expense_usage.values() if e['expense'].expense_type == 'b2c' and e['order_count'] > 0])
        ])
    if both_expense_cost > 0:
        both_avg = both_expense_cost / both_count if both_count > 0 else 0
        type_data.append([
            "Both B2B & B2C", 
            float(both_expense_cost), 
            f"{both_percentage:.1f}%",
            both_count,
            float(both_avg),
            len([e for e in expense_usage.values() if e['expense'].expense_type == 'both' and e['order_count'] > 0])
        ])
    
    exporter.add_sheet(SheetConfig(
        name="Type Distribution",
        headers=[
            "Expense Type", "Total Cost", "% of Total", 
            "Order Count", "Avg per Order", "Active Expenses"
        ],
        data=type_data,
        column_widths=[25, 15, 12, 12, 15, 15]
    ))
    
    # Sheet 6: Daily Expense Trend
    daily_expenses = orders_base.annotate(
        date=TruncDate('created_at')
    ).values('date').annotate(
        order_count=Count('id')
    ).order_by('date')
    
    # Calculate daily expense costs
    daily_trend_data = []
    for daily_item in daily_expenses:
        date = daily_item['date']
        daily_orders = orders_base.filter(created_at__date=date)
        
        daily_cost = Decimal('0')
        daily_b2b = Decimal('0')
        daily_b2c = Decimal('0')
        
        for expense in expenses:
            products_with_expense = expense.products.all()
            related_orders = daily_orders.filter(product__in=products_with_expense)
            order_count = related_orders.count()
            expense_cost = Decimal('0')
            for order in related_orders:
                expense_cost += expense.price_for_original + (expense.price_for_copy * (order.copy_number or 0))
            daily_cost += expense_cost
            
            if expense.expense_type == 'b2b':
                daily_b2b += expense_cost
            elif expense.expense_type == 'b2c':
                daily_b2c += expense_cost
        
        if daily_cost > 0:
            daily_trend_data.append([
                date,
                daily_item['order_count'],
                float(daily_cost),
                float(daily_cost / daily_item['order_count']) if daily_item['order_count'] > 0 else 0,
                float(daily_b2b),
                float(daily_b2c),
            ])
    
    exporter.add_sheet(SheetConfig(
        name="Daily Expense Trend",
        headers=[
            "Date", "Orders", "Total Expense Cost", 
            "Avg per Order", "B2B Cost", "B2C Cost"
        ],
        data=daily_trend_data,
        column_widths=[15, 10, 18, 15, 15, 15]
    ))
    
    # Generate filename
    filename = f"expense_analytics_{date_from.strftime('%Y%m%d')}_to_{date_to.strftime('%Y%m%d')}.xlsx"
    
    return exporter.generate_response(filename)
